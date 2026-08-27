from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[2]

WORKBOOK_PATH = (
    ROOT
    / "data/raw/reference"
    / "estim-pop-nreg-sexe-aq-1975-2026.xlsx"
)

CROSSWALK_PATH = (
    ROOT
    / "data/metadata/insee_region_crosswalk.csv"
)

SEX_CODES = {
    "Hommes": "1",
    "Femmes": "2",
}

EXPECTED_SEX_BLOCKS = {
    "Ensemble",
    "Hommes",
    "Femmes",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build harmonised INSEE population denominators "
            "for selected study years."
        )
    )
    parser.add_argument(
        "--mode",
        choices=("pilot", "full"),
        default="pilot",
        help=(
            "Build pilot outputs for 2019 and 2025 "
            "or full outputs for 2019 through 2025."
        ),
    )
    parser.add_argument(
        "--years",
        nargs="+",
        type=int,
        default=None,
        help=(
            "Study years to process in pilot mode; "
            "defaults to 2019 and 2025."
        ),
    )
    args = parser.parse_args()

    if args.mode == "full" and args.years is not None:
        parser.error(
            "--years cannot be combined with --mode full"
        )

    return args


def relative_path(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def normalize_text(value: object) -> str:
    if value is None:
        return ""

    text = str(value).replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"(\d)à", r"\1 à", text)

    return text


def population_status(source_text: str) -> str:
    lowered = source_text.lower()

    if "provisoire" in lowered:
        return "provisional"
    if "précoce" in lowered:
        return "early"
    if "définitif" in lowered:
        return "definitive"

    return "not_specified"


def age_lower_bound(age_label: str) -> int:
    match = re.match(r"^(\d+)", age_label)

    if match is None:
        raise ValueError(
            f"Unable to parse age label: {age_label}"
        )

    return int(match.group(1))


def open_medic_age_code(lower_bound: int) -> str:
    if lower_bound <= 19:
        return "0"
    if lower_bound <= 59:
        return "20"
    return "60"


def normalize_population(
    value: object,
    location: str,
) -> int:
    if value is None:
        raise ValueError(
            f"Missing population value at {location}"
        )

    try:
        numeric = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(
            f"Non-numeric population value at {location}: "
            f"{value!r}"
        ) from error

    if not numeric.is_integer():
        raise ValueError(
            f"Non-integer population value at {location}: "
            f"{value!r}"
        )

    integer = int(numeric)

    if integer < 0:
        raise ValueError(
            f"Negative population value at {location}: "
            f"{integer}"
        )

    return integer


def find_source_text(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> str:
    for row_number in range(1, sheet.max_row + 1):
        value = normalize_text(
            sheet.cell(
                row=row_number,
                column=1,
            ).value
        )
        if value.startswith("Source :"):
            return value

    raise ValueError(
        f"No source note found in sheet {sheet.title}"
    )


def parse_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    reference_year: int,
    crosswalk: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, object]]:
    expected_title = (
        "Estimation de population au 1 janvier"
    )
    title = normalize_text(
        sheet.cell(row=1, column=1).value
    ).replace("1er", "1")

    if expected_title not in title:
        raise ValueError(
            f"Unexpected title in sheet {sheet.title}: "
            f"{title}"
        )

    year_label = normalize_text(
        sheet.cell(row=2, column=1).value
    )

    if year_label != f"Année {reference_year}":
        raise ValueError(
            f"Unexpected year label in sheet "
            f"{sheet.title}: {year_label}"
        )

    source_text = find_source_text(sheet)
    status = population_status(source_text)

    column_specs: list[dict[str, object]] = []
    current_sex = ""

    for column_number in range(
        2,
        sheet.max_column + 1,
    ):
        block_value = normalize_text(
            sheet.cell(
                row=4,
                column=column_number,
            ).value
        )

        if block_value:
            current_sex = block_value

        age_label = normalize_text(
            sheet.cell(
                row=5,
                column=column_number,
            ).value
        )

        if not age_label:
            continue

        if current_sex not in EXPECTED_SEX_BLOCKS:
            raise ValueError(
                f"Unexpected sex block in sheet "
                f"{sheet.title}: {current_sex}"
            )

        column_specs.append(
            {
                "column": column_number,
                "sex_label": current_sex,
                "age_label": age_label,
            }
        )

    block_counts = pd.Series(
        [
            specification["sex_label"]
            for specification in column_specs
        ]
    ).value_counts()

    for sex_label in EXPECTED_SEX_BLOCKS:
        if block_counts.get(sex_label, 0) != 21:
            raise ValueError(
                f"Unexpected number of columns for "
                f"{sex_label} in sheet {sheet.title}"
            )

    expected_regions = set(
        crosswalk["insee_region_name"]
    )
    region_rows: dict[str, int] = {}

    for row_number in range(
        1,
        sheet.max_row + 1,
    ):
        region_name = normalize_text(
            sheet.cell(
                row=row_number,
                column=1,
            ).value
        )

        if region_name in expected_regions:
            region_rows[region_name] = row_number

    if set(region_rows) != expected_regions:
        missing = sorted(
            expected_regions.difference(region_rows)
        )
        raise ValueError(
            f"Missing regions in sheet {sheet.title}: "
            f"{missing}"
        )

    records: list[dict[str, object]] = []

    for region_name, row_number in region_rows.items():
        for specification in column_specs:
            column_number = int(
                specification["column"]
            )
            sex_label = str(
                specification["sex_label"]
            )
            age_label = str(
                specification["age_label"]
            )

            value = normalize_population(
                sheet.cell(
                    row=row_number,
                    column=column_number,
                ).value,
                (
                    f"{sheet.title}!R{row_number}"
                    f"C{column_number}"
                ),
            )

            records.append(
                {
                    "reference_year": reference_year,
                    "insee_region_name": region_name,
                    "sex_label": sex_label,
                    "age_label": age_label,
                    "population": value,
                }
            )

    parsed = pd.DataFrame(records)

    expected_rows = (
        len(expected_regions)
        * len(EXPECTED_SEX_BLOCKS)
        * 21
    )

    if len(parsed) != expected_rows:
        raise ValueError(
            f"Unexpected parsed row count in sheet "
            f"{sheet.title}: {len(parsed)}"
        )

    pivot = parsed.pivot(
        index=["insee_region_name", "age_label"],
        columns="sex_label",
        values="population",
    )

    if not (
        pivot["Ensemble"]
        == pivot["Hommes"] + pivot["Femmes"]
    ).all():
        raise ValueError(
            f"Sex totals do not reconcile in "
            f"sheet {sheet.title}"
        )

    age_rows = parsed.loc[
        parsed["age_label"].ne("Total")
    ]

    calculated_totals = (
        age_rows.groupby(
            ["insee_region_name", "sex_label"],
            as_index=False,
        )["population"]
        .sum()
        .rename(
            columns={
                "population": "calculated_total"
            }
        )
    )

    reported_totals = (
        parsed.loc[
            parsed["age_label"].eq("Total"),
            [
                "insee_region_name",
                "sex_label",
                "population",
            ],
        ]
        .rename(
            columns={
                "population": "reported_total"
            }
        )
    )

    total_check = calculated_totals.merge(
        reported_totals,
        on=["insee_region_name", "sex_label"],
        validate="one_to_one",
    )

    if not (
        total_check["calculated_total"]
        == total_check["reported_total"]
    ).all():
        raise ValueError(
            f"Age totals do not reconcile in "
            f"sheet {sheet.title}"
        )

    selected = parsed.loc[
        parsed["sex_label"].isin(SEX_CODES)
        & parsed["age_label"].ne("Total")
    ].copy()

    selected["sex_code"] = selected[
        "sex_label"
    ].map(SEX_CODES)

    selected["age_lower"] = selected[
        "age_label"
    ].map(age_lower_bound)

    selected["age_code"] = selected[
        "age_lower"
    ].map(open_medic_age_code)

    selected["population_status"] = status
    selected["source_sheet"] = sheet.title
    selected["source_file"] = relative_path(
        WORKBOOK_PATH
    )

    selected = selected.merge(
        crosswalk,
        on="insee_region_name",
        how="left",
        validate="many_to_one",
    )

    if selected[
        "analysis_region_code"
    ].isna().any():
        raise ValueError(
            f"Unmapped INSEE regions in sheet "
            f"{sheet.title}"
        )

    group_columns = [
        "reference_year",
        "analysis_region_code",
        "analysis_region_name",
        "analysis_scope",
        "sex_code",
        "age_label",
        "age_lower",
        "age_code",
        "population_status",
        "source_sheet",
        "source_file",
    ]

    harmonised = (
        selected.groupby(
            group_columns,
            as_index=False,
            dropna=False,
        )["population"]
        .sum()
        .rename(
            columns={
                "population": "population_january1"
            }
        )
    )

    expected_analysis_regions = (
        crosswalk["analysis_region_code"].nunique()
    )
    expected_harmonised_rows = (
        expected_analysis_regions
        * 2
        * 20
    )

    if len(harmonised) != expected_harmonised_rows:
        raise ValueError(
            f"Unexpected harmonised row count in "
            f"sheet {sheet.title}: {len(harmonised)}"
        )

    inventory = {
        "reference_year": reference_year,
        "source_sheet": sheet.title,
        "population_status": status,
        "source_note": source_text,
        "source_region_rows": len(region_rows),
        "validated_source_cells": len(parsed),
        "harmonised_rows": len(harmonised),
        "sex_totals_reconciled": True,
        "age_totals_reconciled": True,
        "source_file": relative_path(
            WORKBOOK_PATH
        ),
    }

    return harmonised, inventory


def build_annual_denominators(
    january_population: pd.DataFrame,
    study_years: list[int],
) -> pd.DataFrame:
    strata_columns = [
        "reference_year",
        "analysis_region_code",
        "analysis_region_name",
        "analysis_scope",
        "sex_code",
        "age_code",
        "population_status",
    ]

    strata = (
        january_population.groupby(
            strata_columns,
            as_index=False,
        )["population_january1"]
        .sum()
    )

    annual_frames: list[pd.DataFrame] = []

    join_columns = [
        "analysis_region_code",
        "analysis_region_name",
        "analysis_scope",
        "sex_code",
        "age_code",
    ]

    for study_year in study_years:
        start = strata.loc[
            strata["reference_year"].eq(study_year)
        ].copy()

        end = strata.loc[
            strata["reference_year"].eq(
                study_year + 1
            )
        ].copy()

        start = start.drop(
            columns=["reference_year"]
        ).rename(
            columns={
                "population_january1":
                    "population_start",
                "population_status":
                    "population_start_status",
            }
        )

        end = end.drop(
            columns=["reference_year"]
        ).rename(
            columns={
                "population_january1":
                    "population_end",
                "population_status":
                    "population_end_status",
            }
        )

        annual = start.merge(
            end,
            on=join_columns,
            validate="one_to_one",
        )

        annual.insert(0, "study_year", study_year)
        annual["population_average"] = (
            annual["population_start"]
            + annual["population_end"]
        ) / 2

        annual_frames.append(annual)

    result = pd.concat(
        annual_frames,
        ignore_index=True,
    )

    expected_rows = (
        len(study_years)
        * result["analysis_region_code"].nunique()
        * 2
        * 3
    )

    if len(result) != expected_rows:
        raise ValueError(
            f"Unexpected annual denominator row count: "
            f"{len(result)}"
        )

    return result.sort_values(
        [
            "study_year",
            "analysis_region_code",
            "age_code",
            "sex_code",
        ]
    ).reset_index(drop=True)


def main() -> None:
    args = parse_args()
    if args.mode == "full":
        study_years = list(range(2019, 2026))
    else:
        study_years = sorted(
            set(args.years or [2019, 2025])
        )

    reference_years = sorted(
        {
            year
            for study_year in study_years
            for year in (
                study_year,
                study_year + 1,
            )
        }
    )

    crosswalk = pd.read_csv(
        CROSSWALK_PATH,
        dtype=str,
        keep_default_na=False,
    )

    if not crosswalk[
        "insee_region_name"
    ].is_unique:
        raise ValueError(
            "Duplicate INSEE region names in crosswalk."
        )

    workbook = openpyxl.load_workbook(
        WORKBOOK_PATH,
        read_only=True,
        data_only=True,
    )

    frames: list[pd.DataFrame] = []
    inventory_rows: list[dict[str, object]] = []

    try:
        for reference_year in reference_years:
            sheet_name = str(reference_year)

            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Missing INSEE sheet: {sheet_name}"
                )

            harmonised, inventory = parse_sheet(
                workbook[sheet_name],
                reference_year,
                crosswalk,
            )
            frames.append(harmonised)
            inventory_rows.append(inventory)
    finally:
        workbook.close()

    january_population = pd.concat(
        frames,
        ignore_index=True,
    ).sort_values(
        [
            "reference_year",
            "analysis_region_code",
            "age_lower",
            "sex_code",
        ]
    ).reset_index(drop=True)

    annual_denominators = build_annual_denominators(
        january_population,
        study_years,
    )

    output_dir = (
        ROOT / "data/interim/insee" / args.mode
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    filename_suffix = (
        "_pilot" if args.mode == "pilot" else ""
    )

    january_path = (
        output_dir
        / f"insee_population_january1{filename_suffix}.parquet"
    )
    annual_path = (
        output_dir
        / f"insee_annual_denominators{filename_suffix}.parquet"
    )

    january_population.to_parquet(
        january_path,
        index=False,
    )
    annual_denominators.to_parquet(
        annual_path,
        index=False,
    )

    if 2025 in study_years:
        standard = annual_denominators.loc[
            annual_denominators[
                "study_year"
            ].eq(2025)
            & annual_denominators[
                "analysis_region_code"
            ].eq("FR")
        ].copy()

        if len(standard) != 6:
            raise ValueError(
                "Expected six 2025 standard strata."
            )

        standard["standard_weight"] = (
            standard["population_average"]
            / standard["population_average"].sum()
        )

        if abs(
            standard["standard_weight"].sum() - 1
        ) > 1e-12:
            raise ValueError(
                "Standard weights do not sum to one."
            )

        standard_path = (
            output_dir
            / f"insee_standard_population_2025{filename_suffix}.parquet"
        )
        standard.to_parquet(
            standard_path,
            index=False,
        )
        print(
            f"created {relative_path(standard_path)} "
            f"({len(standard)} rows)"
        )

    inventory = pd.DataFrame(
        inventory_rows
    ).sort_values("reference_year")

    inventory["january_output_file"] = (
        relative_path(january_path)
    )
    inventory["annual_output_file"] = (
        relative_path(annual_path)
    )

    inventory_name = (
        "insee_pilot_inventory.csv"
        if args.mode == "pilot"
        else "insee_inventory.csv"
    )
    inventory_path = (
        ROOT / "data/metadata" / inventory_name
    )
    inventory.to_csv(
        inventory_path,
        index=False,
        lineterminator="\n",
    )

    print(
        f"created {relative_path(january_path)} "
        f"({len(january_population)} rows)"
    )
    print(
        f"created {relative_path(annual_path)} "
        f"({len(annual_denominators)} rows)"
    )
    print(
        f"Inventory: {relative_path(inventory_path)}"
    )
    print(
        "Study years: "
        + ", ".join(map(str, study_years))
    )
    print(
        "Reference years: "
        + ", ".join(map(str, reference_years))
    )


if __name__ == "__main__":
    main()
