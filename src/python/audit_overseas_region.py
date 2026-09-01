"""Audit the Open Medic overseas regional grouping used in the study.

This script independently checks the numerator, denominator, age-sex
standardisation, source-file provenance, and the influence of Mayotte on the
combined overseas rate. It writes machine-readable QC tables and a concise
Markdown summary under ``output/tables``.
"""

from __future__ import annotations

import hashlib
import re
from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[2]
STUDY_YEARS = tuple(range(2020, 2026))
REFERENCE_YEARS = tuple(range(2020, 2027))
AGE_CODES = ("0", "20", "60")
SEX_CODES = ("1", "2")
OVERSEAS_CODE = "5"
OVERSEAS_COMPONENTS = (
    "Guadeloupe",
    "Martinique",
    "Guyane",
    "La Réunion",
    "Mayotte",
)
OVERSEAS_COMPONENTS_WITHOUT_MAYOTTE = tuple(
    name for name in OVERSEAS_COMPONENTS if name != "Mayotte"
)

WORKBOOK_PATH = (
    ROOT
    / "data/raw/reference"
    / "estim-pop-nreg-sexe-aq-1975-2026.xlsx"
)
REGIONAL_ANNUAL_PATH = (
    ROOT / "data/processed/regional_annual.parquet"
)
REGIONAL_CELLS_PATH = (
    ROOT / "data/processed/regional_age_sex_rates.parquet"
)
REGIONAL_STANDARDISED_PATH = (
    ROOT / "data/processed/regional_standardised_rates.parquet"
)
OPEN_MEDIC_REGION_PATH = (
    ROOT
    / "data/interim/open_medic/full"
    / "open_medic_atc4_region.parquet"
)
OPEN_MEDIC_CELLS_PATH = (
    ROOT
    / "data/interim/open_medic/full"
    / "open_medic_atc4_age_sex_region.parquet"
)
DENOMINATOR_PATH = (
    ROOT
    / "data/interim/insee/full"
    / "insee_annual_denominators.parquet"
)
DOWNLOAD_MANIFEST_PATH = (
    ROOT / "data/metadata/download_manifest.csv"
)
SOURCE_CATALOG_PATH = ROOT / "data/metadata/source_catalog.csv"
OUTPUT_DIRECTORY = ROOT / "output/tables"


def normalize_text(value: object) -> str:
    """Normalize workbook labels without changing their meaning."""
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def age_code_from_label(label: str) -> str:
    """Map an INSEE five-year age label to the Open Medic age bands."""
    match = re.match(r"^(\d+)", label)
    if match is None:
        raise ValueError(f"Unable to parse age label: {label!r}")
    lower = int(match.group(1))
    if lower <= 19:
        return "0"
    if lower <= 59:
        return "20"
    return "60"


def require_columns(
    frame: pd.DataFrame,
    columns: set[str],
    label: str,
) -> None:
    """Raise a readable error when an expected schema is incomplete."""
    missing = sorted(columns.difference(frame.columns))
    if missing:
        raise ValueError(f"Missing columns in {label}: {missing}")


def sha256_file(path: Path) -> str:
    """Calculate a source-file checksum without loading it into memory."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def locate_year_sheets(
    workbook: openpyxl.Workbook,
) -> dict[int, openpyxl.worksheet.worksheet.Worksheet]:
    """Locate annual sheets from their internal year labels."""
    located: dict[int, openpyxl.worksheet.worksheet.Worksheet] = {}
    for sheet in workbook.worksheets:
        year_label = normalize_text(sheet.cell(row=2, column=1).value)
        match = re.fullmatch(r"Année (\d{4})", year_label)
        if match:
            located[int(match.group(1))] = sheet

    missing = sorted(set(REFERENCE_YEARS).difference(located))
    if missing:
        raise ValueError(f"Missing INSEE workbook sheets: {missing}")
    return located


def parse_insee_overseas_strata() -> pd.DataFrame:
    """Read DOM and its five components directly from the official workbook."""
    workbook = openpyxl.load_workbook(
        WORKBOOK_PATH,
        read_only=True,
        data_only=True,
    )
    sheets = locate_year_sheets(workbook)
    target_regions = {"DOM", *OVERSEAS_COMPONENTS}
    sex_codes = {"Hommes": "1", "Femmes": "2"}
    records: list[dict[str, object]] = []

    for reference_year in REFERENCE_YEARS:
        sheet = sheets[reference_year]
        region_rows: dict[str, int] = {}
        for row_number in range(1, sheet.max_row + 1):
            region_name = normalize_text(
                sheet.cell(row=row_number, column=1).value
            )
            if region_name in target_regions:
                region_rows[region_name] = row_number

        missing_regions = sorted(target_regions.difference(region_rows))
        if missing_regions:
            raise ValueError(
                f"Missing regions in INSEE {reference_year}: "
                f"{missing_regions}"
            )

        current_sex = ""
        column_specs: list[tuple[int, str, str]] = []
        for column_number in range(2, sheet.max_column + 1):
            block = normalize_text(
                sheet.cell(row=4, column=column_number).value
            )
            if block:
                current_sex = block
            age_label = normalize_text(
                sheet.cell(row=5, column=column_number).value
            )
            if (
                current_sex in sex_codes
                and age_label
                and age_label != "Total"
            ):
                column_specs.append(
                    (
                        column_number,
                        sex_codes[current_sex],
                        age_code_from_label(age_label),
                    )
                )

        for region_name, row_number in region_rows.items():
            for column_number, sex_code, age_code in column_specs:
                value = sheet.cell(
                    row=row_number,
                    column=column_number,
                ).value
                if value is None:
                    raise ValueError(
                        "Missing INSEE population at "
                        f"{sheet.title}!R{row_number}C{column_number}"
                    )
                records.append(
                    {
                        "reference_year": reference_year,
                        "source_region": region_name,
                        "age_code": age_code,
                        "sex_code": sex_code,
                        "population": int(value),
                    }
                )

    workbook.close()
    parsed = pd.DataFrame.from_records(records)
    parsed = (
        parsed.groupby(
            [
                "reference_year",
                "source_region",
                "age_code",
                "sex_code",
            ],
            as_index=False,
        )["population"]
        .sum()
        .sort_values(
            [
                "reference_year",
                "source_region",
                "age_code",
                "sex_code",
            ]
        )
    )
    expected_rows = len(REFERENCE_YEARS) * len(target_regions) * 6
    if len(parsed) != expected_rows:
        raise ValueError(
            f"Unexpected parsed INSEE rows: {len(parsed)}; "
            f"expected {expected_rows}"
        )
    return parsed


def build_denominator_scope_audit(
    parsed: pd.DataFrame,
) -> pd.DataFrame:
    """Compare the direct DOM aggregate with four- and five-DROM sums."""
    keys = ["reference_year", "age_code", "sex_code"]
    direct = (
        parsed.loc[parsed["source_region"] == "DOM", keys + ["population"]]
        .rename(columns={"population": "direct_dom_population"})
    )
    five = (
        parsed.loc[parsed["source_region"].isin(OVERSEAS_COMPONENTS)]
        .groupby(keys, as_index=False)["population"]
        .sum()
        .rename(columns={"population": "five_drom_population"})
    )
    four = (
        parsed.loc[
            parsed["source_region"].isin(
                OVERSEAS_COMPONENTS_WITHOUT_MAYOTTE
            )
        ]
        .groupby(keys, as_index=False)["population"]
        .sum()
        .rename(columns={"population": "four_drom_population"})
    )
    mayotte = (
        parsed.loc[
            parsed["source_region"] == "Mayotte",
            keys + ["population"],
        ]
        .rename(columns={"population": "mayotte_population"})
    )
    audit = direct.merge(five, on=keys, validate="one_to_one")
    audit = audit.merge(four, on=keys, validate="one_to_one")
    audit = audit.merge(mayotte, on=keys, validate="one_to_one")
    audit["direct_minus_five_drom"] = (
        audit["direct_dom_population"] - audit["five_drom_population"]
    )
    audit["mayotte_share_of_dom_pct"] = (
        audit["mayotte_population"]
        / audit["direct_dom_population"]
        * 100
    )
    audit["dom_matches_five_drom"] = (
        audit["direct_minus_five_drom"] == 0
    )
    return audit.sort_values(keys)


def annual_average_denominators(
    scope_audit: pd.DataFrame,
) -> pd.DataFrame:
    """Calculate average annual DOM denominators for sensitivity checks."""
    totals = (
        scope_audit.groupby("reference_year", as_index=False)
        .agg(
            direct_dom_population=("direct_dom_population", "sum"),
            five_drom_population=("five_drom_population", "sum"),
            four_drom_population=("four_drom_population", "sum"),
            mayotte_population=("mayotte_population", "sum"),
        )
        .sort_values("reference_year")
    )
    current = totals.rename(
        columns={
            "reference_year": "study_year",
            "direct_dom_population": "direct_dom_start",
            "five_drom_population": "five_drom_start",
            "four_drom_population": "four_drom_start",
            "mayotte_population": "mayotte_start",
        }
    )
    following = totals.assign(
        study_year=totals["reference_year"] - 1
    ).rename(
        columns={
            "direct_dom_population": "direct_dom_end",
            "five_drom_population": "five_drom_end",
            "four_drom_population": "four_drom_end",
            "mayotte_population": "mayotte_end",
        }
    )
    result = current.merge(
        following[
            [
                "study_year",
                "direct_dom_end",
                "five_drom_end",
                "four_drom_end",
                "mayotte_end",
            ]
        ],
        on="study_year",
        validate="one_to_one",
    )
    result = result.loc[result["study_year"].isin(STUDY_YEARS)].copy()
    for prefix in ("direct_dom", "five_drom", "four_drom", "mayotte"):
        result[f"{prefix}_average"] = (
            result[f"{prefix}_start"] + result[f"{prefix}_end"]
        ) / 2
    return result.sort_values("study_year")


def audit_open_medic_reconciliation() -> pd.DataFrame:
    """Reconcile direct and stratified Open Medic overseas numerators."""
    direct = pd.read_parquet(OPEN_MEDIC_REGION_PATH)
    cells = pd.read_parquet(OPEN_MEDIC_CELLS_PATH)
    for frame in (direct, cells):
        frame["region_code"] = frame["region_code"].astype("string")
    direct = direct.loc[
        (direct["region_code"] == OVERSEAS_CODE)
        & direct["year"].isin(STUDY_YEARS)
    ]
    stratified = (
        cells.loc[
            (cells["region_code"] == OVERSEAS_CODE)
            & cells["year"].isin(STUDY_YEARS)
        ]
        .groupby(["year", "region_code"], as_index=False)
        .agg(
            stratified_beneficiaries=("beneficiaries", "sum"),
            stratified_boxes=("boxes", "sum"),
            stratified_expenditure_eur=(
                "reimbursed_expenditure_eur",
                "sum",
            ),
            stratified_rows=("beneficiaries", "size"),
        )
    )
    reconciliation = direct[
        [
            "year",
            "region_code",
            "beneficiaries",
            "boxes",
            "reimbursed_expenditure_eur",
            "source_file",
        ]
    ].merge(
        stratified,
        on=["year", "region_code"],
        validate="one_to_one",
    )
    reconciliation = reconciliation.rename(
        columns={
            "beneficiaries": "direct_beneficiaries",
            "boxes": "direct_boxes",
            "reimbursed_expenditure_eur": "direct_expenditure_eur",
        }
    )
    reconciliation["beneficiary_difference"] = (
        reconciliation["stratified_beneficiaries"]
        - reconciliation["direct_beneficiaries"]
    )
    reconciliation["boxes_difference"] = (
        reconciliation["stratified_boxes"]
        - reconciliation["direct_boxes"]
    )
    reconciliation["expenditure_difference_eur"] = (
        reconciliation["stratified_expenditure_eur"]
        - reconciliation["direct_expenditure_eur"]
    )
    return reconciliation.sort_values("year")


def audit_annual_rates(
    average_denominators: pd.DataFrame,
) -> pd.DataFrame:
    """Recalculate annual overseas rates and denominator sensitivities."""
    annual = pd.read_parquet(REGIONAL_ANNUAL_PATH)
    annual["region_code"] = annual["region_code"].astype("string")
    overseas = annual.loc[
        annual["region_code"] == OVERSEAS_CODE
    ].copy()
    metropolitan = annual.loc[
        annual["region_code"] != OVERSEAS_CODE
    ].copy()
    comparison = (
        metropolitan.groupby("study_year", as_index=False)
        .agg(
            metropolitan_median_crude_rate=(
                "beneficiary_rate_per_100000",
                "median",
            ),
            metropolitan_max_crude_rate=(
                "beneficiary_rate_per_100000",
                "max",
            ),
        )
    )
    audit = overseas.merge(
        comparison,
        on="study_year",
        validate="one_to_one",
    ).merge(
        average_denominators,
        on="study_year",
        validate="one_to_one",
    )
    audit["recalculated_crude_rate_per_100000"] = (
        audit["beneficiaries"] / audit["population_average"] * 100000
    )
    audit["rate_recalculation_difference"] = (
        audit["recalculated_crude_rate_per_100000"]
        - audit["beneficiary_rate_per_100000"]
    )
    audit["processed_minus_direct_dom_population"] = (
        audit["population_average"] - audit["direct_dom_average"]
    )
    audit["rate_using_direct_dom_per_100000"] = (
        audit["beneficiaries"] / audit["direct_dom_average"] * 100000
    )
    audit["rate_using_five_drom_per_100000"] = (
        audit["beneficiaries"] / audit["five_drom_average"] * 100000
    )
    audit["rate_using_four_drom_per_100000"] = (
        audit["beneficiaries"] / audit["four_drom_average"] * 100000
    )
    audit["overseas_to_metro_median_rate_ratio"] = (
        audit["beneficiary_rate_per_100000"]
        / audit["metropolitan_median_crude_rate"]
    )
    audit["population_needed_for_metro_median"] = (
        audit["beneficiaries"]
        / audit["metropolitan_median_crude_rate"]
        * 100000
    )
    audit["required_population_increase_pct"] = (
        audit["population_needed_for_metro_median"]
        / audit["population_average"]
        - 1
    ) * 100
    return audit.sort_values("study_year")


def audit_age_sex_pattern() -> pd.DataFrame:
    """Compare each overseas age-sex cell with its metropolitan median."""
    cells = pd.read_parquet(REGIONAL_CELLS_PATH)
    cells["region_code"] = cells["region_code"].astype("string")
    observed = cells.loc[
        cells["beneficiary_rate_per_100000"].notna()
    ].copy()
    metropolitan = observed.loc[
        observed["region_code"] != OVERSEAS_CODE
    ]
    medians = (
        metropolitan.groupby(
            ["study_year", "age_code", "sex_code"],
            as_index=False,
        )
        .agg(
            metropolitan_median_cell_rate=(
                "beneficiary_rate_per_100000",
                "median",
            ),
            metropolitan_max_cell_rate=(
                "beneficiary_rate_per_100000",
                "max",
            ),
        )
    )
    overseas = cells.loc[
        cells["region_code"] == OVERSEAS_CODE
    ].copy()
    result = overseas.merge(
        medians,
        on=["study_year", "age_code", "sex_code"],
        how="left",
        validate="one_to_one",
    )
    result["overseas_to_metro_median_cell_rate_ratio"] = (
        result["beneficiary_rate_per_100000"]
        / result["metropolitan_median_cell_rate"]
    )
    result["overseas_exceeds_metro_max"] = (
        result["beneficiary_rate_per_100000"]
        > result["metropolitan_max_cell_rate"]
    )
    return result.sort_values(
        ["study_year", "age_code", "sex_code"]
    )


def audit_standardisation(
    age_sex_audit: pd.DataFrame,
) -> pd.DataFrame:
    """Independently reproduce overseas direct standardisation."""
    denominators = pd.read_parquet(DENOMINATOR_PATH)
    denominators["analysis_region_code"] = denominators[
        "analysis_region_code"
    ].astype("string")
    standard = denominators.loc[
        (denominators["study_year"] == 2025)
        & (denominators["analysis_region_code"] == "FR")
        & denominators["age_code"].astype("string").isin(AGE_CODES)
        & denominators["sex_code"].astype("string").isin(SEX_CODES)
    ].copy()
    standard["age_code"] = standard["age_code"].astype("string")
    standard["sex_code"] = standard["sex_code"].astype("string")
    standard["standard_weight"] = (
        standard["population_average"]
        / standard["population_average"].sum()
    )
    require_columns(
        standard,
        {"age_code", "sex_code", "standard_weight"},
        "standard population",
    )
    if len(standard) != 6:
        raise ValueError(f"Expected 6 standard strata; found {len(standard)}")

    cells = age_sex_audit.drop(
        columns=["standard_population", "standard_weight"],
        errors="ignore",
    ).merge(
        standard[["age_code", "sex_code", "standard_weight"]],
        on=["age_code", "sex_code"],
        validate="many_to_one",
    )
    reproduced = (
        cells.assign(
            weighted_point=(
                cells["beneficiary_rate_per_100000"]
                * cells["standard_weight"]
            ),
            weighted_lower=(
                cells["beneficiary_rate_lower_per_100000"]
                * cells["standard_weight"]
            ),
            weighted_upper=(
                cells["beneficiary_rate_upper_per_100000"]
                * cells["standard_weight"]
            ),
        )
        .groupby("study_year", as_index=False)
        .agg(
            reproduced_point=("weighted_point", lambda x: x.sum(min_count=6)),
            reproduced_lower=("weighted_lower", "sum"),
            reproduced_upper=("weighted_upper", "sum"),
            strata=("standard_weight", "size"),
        )
    )
    published = pd.read_parquet(REGIONAL_STANDARDISED_PATH)
    published["region_code"] = published["region_code"].astype("string")
    published = published.loc[
        published["region_code"] == OVERSEAS_CODE
    ]
    result = published.merge(
        reproduced,
        on="study_year",
        validate="one_to_one",
    )
    result["point_difference"] = (
        result["reproduced_point"]
        - result["standardised_rate_per_100000"]
    )
    result["lower_difference"] = (
        result["reproduced_lower"]
        - result["standardised_rate_lower_per_100000"]
    )
    result["upper_difference"] = (
        result["reproduced_upper"]
        - result["standardised_rate_upper_per_100000"]
    )
    return result.sort_values("study_year")


def audit_source_provenance() -> pd.DataFrame:
    """Confirm that the local 2025 regional files match their manifest."""
    manifest = pd.read_csv(
        DOWNLOAD_MANIFEST_PATH,
        dtype="string",
    )
    require_columns(
        manifest,
        {
            "year",
            "atc_level",
            "variant",
            "local_path",
            "downloaded_at_utc",
            "sha256",
        },
        "download manifest",
    )
    relevant = manifest.loc[
        (manifest["year"] == "2025")
        & (manifest["atc_level"].str.upper() == "ATC4")
        & manifest["variant"].isin(["region", "age_sex_region"])
    ].copy()
    if len(relevant) != 2:
        raise ValueError(
            "Expected two 2025 ATC4 regional manifest rows; "
            f"found {len(relevant)}"
        )
    relevant["file_exists"] = relevant["local_path"].map(
        lambda value: (ROOT / str(value)).is_file()
    )
    relevant["calculated_sha256"] = relevant.apply(
        lambda row: sha256_file(ROOT / str(row["local_path"]))
        if bool(row["file_exists"])
        else pd.NA,
        axis=1,
    )
    relevant["checksum_matches"] = (
        relevant["calculated_sha256"] == relevant["sha256"]
    )
    relevant["downloaded_after_2025_correction"] = (
        pd.to_datetime(relevant["downloaded_at_utc"], utc=True)
        >= pd.Timestamp("2026-07-10", tz="UTC")
    )

    catalog = pd.read_csv(SOURCE_CATALOG_PATH, dtype="string")
    notes = " ".join(
        catalog.loc[
            (catalog["year"] == "2025")
            & (catalog["atc_level"].str.upper() == "ATC4"),
            "notes",
        ].dropna()
    )
    relevant["catalog_mentions_corrected_regional_data"] = bool(
        re.search(r"correct", notes, flags=re.IGNORECASE)
    )
    return relevant.sort_values("variant")


def write_summary(
    denominator_scope: pd.DataFrame,
    reconciliation: pd.DataFrame,
    annual: pd.DataFrame,
    age_sex: pd.DataFrame,
    standardisation: pd.DataFrame,
    provenance: pd.DataFrame,
) -> Path:
    """Write a concise human-readable audit report."""
    denominator_match = bool(denominator_scope["dom_matches_five_drom"].all())
    processed_denominator_match = bool(
        annual["processed_minus_direct_dom_population"].abs().max()
        <= 1e-9
    )
    numerator_match = bool(
        (reconciliation["beneficiary_difference"] == 0).all()
        and (reconciliation["boxes_difference"] == 0).all()
        and (reconciliation["expenditure_difference_eur"].abs() <= 0.05).all()
    )
    arithmetic_match = bool(
        annual["rate_recalculation_difference"].abs().max() <= 1e-9
    )
    standardisation_match = bool(
        standardisation[
            ["point_difference", "lower_difference", "upper_difference"]
        ].abs().max().max()
        <= 1e-9
    )
    provenance_match = bool(
        provenance["checksum_matches"].all()
        and provenance["downloaded_after_2025_correction"].all()
    )
    adult_cells = age_sex.loc[
        age_sex["age_code"].isin(["20", "60"])
        & age_sex["beneficiary_rate_per_100000"].notna()
    ]
    adult_ratio_median = adult_cells[
        "overseas_to_metro_median_cell_rate_ratio"
    ].median()
    required_population_increase_2025 = annual.loc[
        annual["study_year"] == 2025,
        "required_population_increase_pct",
    ].iloc[0]
    mayotte_share_2025 = denominator_scope.loc[
        denominator_scope["reference_year"] == 2025,
        "mayotte_population",
    ].sum() / denominator_scope.loc[
        denominator_scope["reference_year"] == 2025,
        "direct_dom_population",
    ].sum() * 100

    overall = all(
        [
            denominator_match,
            processed_denominator_match,
            numerator_match,
            arithmetic_match,
            standardisation_match,
            provenance_match,
        ]
    )
    conclusion = (
        "The audit found no computational or file-version evidence that the "
        "high overseas rate is an artefact. It should nevertheless remain "
        "explicitly qualified because Open Medic exposes code 5 only as a "
        "combined grouping and because health-insurance beneficiaries and "
        "resident-population denominators are not identical populations."
        if overall
        else
        "At least one audit check failed. Regional overseas results should "
        "not be interpreted or frozen until the failed check is resolved."
    )
    lines = [
        "# QC 07 — Overseas regional grouping audit",
        "",
        f"**Overall computational status:** {'PASS' if overall else 'REVIEW'}",
        "",
        "## Core checks",
        "",
        f"- Direct INSEE DOM equals the sum of the five DROM in every "
        f"year-age-sex cell: **{denominator_match}**.",
        f"- Processed annual denominators equal the direct INSEE DOM "
        f"aggregate: **{processed_denominator_match}**.",
        f"- Open Medic direct and age-sex-stratified numerators reconcile: "
        f"**{numerator_match}**.",
        f"- Published crude rates reproduce from beneficiary counts and "
        f"denominators: **{arithmetic_match}**.",
        f"- Published standardised rates reproduce independently: "
        f"**{standardisation_match}**.",
        f"- Local 2025 regional files match manifest checksums and were "
        f"downloaded after the 10 July 2026 correction: "
        f"**{provenance_match}**.",
        "",
        "## Magnitude checks",
        "",
        f"- Mayotte represents {mayotte_share_2025:.1f}% of the direct DOM "
        "population in the 2025 workbook.",
        f"- Across adult age-sex cells, the median overseas-to-metropolitan "
        f"median rate ratio is {adult_ratio_median:.2f}.",
        f"- To reduce the 2025 crude overseas rate to the metropolitan "
        f"regional median solely through the denominator would require a "
        f"population {required_population_increase_2025:.1f}% larger than "
        "the denominator used.",
        "",
        "## Conclusion",
        "",
        conclusion,
        "",
        "## Official scope notes",
        "",
        "- Open Medic defines `BEN_REG = 5` as *Régions et Départements "
        "d'outre-mer*.",
        "- INSEE states that Mayotte has been included in the France field "
        "since 2014; its post-2017 population values are provisional.",
        "- INSEE states that Guadeloupe estimates exclude Saint-Martin and "
        "Saint-Barthélemy.",
        "",
        "Official sources:",
        "",
        "- https://www.assurance-maladie.ameli.fr/etudes-et-donnees/"
        "open-medic-depenses-beneficiaires-medicaments",
        "- https://www.insee.fr/fr/statistiques/8721456",
    ]
    path = OUTPUT_DIRECTORY / "qc_07_overseas_audit_summary.md"
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def main() -> None:
    """Run the complete overseas audit and write its outputs."""
    required_paths = [
        WORKBOOK_PATH,
        REGIONAL_ANNUAL_PATH,
        REGIONAL_CELLS_PATH,
        REGIONAL_STANDARDISED_PATH,
        OPEN_MEDIC_REGION_PATH,
        OPEN_MEDIC_CELLS_PATH,
        DENOMINATOR_PATH,
        DOWNLOAD_MANIFEST_PATH,
        SOURCE_CATALOG_PATH,
    ]
    missing = [str(path.relative_to(ROOT)) for path in required_paths if not path.is_file()]
    if missing:
        raise FileNotFoundError(f"Missing audit inputs: {missing}")

    OUTPUT_DIRECTORY.mkdir(parents=True, exist_ok=True)
    parsed = parse_insee_overseas_strata()
    denominator_scope = build_denominator_scope_audit(parsed)
    average_denominators = annual_average_denominators(denominator_scope)
    reconciliation = audit_open_medic_reconciliation()
    annual = audit_annual_rates(average_denominators)
    age_sex = audit_age_sex_pattern()
    standardisation = audit_standardisation(age_sex)
    provenance = audit_source_provenance()

    outputs = {
        "qc_07_overseas_denominator_scope_2020_2026.csv": denominator_scope,
        "qc_07_overseas_numerator_reconciliation_2020_2025.csv": reconciliation,
        "qc_07_overseas_annual_rates_2020_2025.csv": annual,
        "qc_07_overseas_age_sex_pattern_2020_2025.csv": age_sex,
        "qc_07_overseas_standardisation_2020_2025.csv": standardisation,
        "qc_07_overseas_source_provenance.csv": provenance,
    }
    for filename, frame in outputs.items():
        path = OUTPUT_DIRECTORY / filename
        frame.to_csv(path, index=False)
        print(f"created {path.relative_to(ROOT)}")

    summary_path = write_summary(
        denominator_scope,
        reconciliation,
        annual,
        age_sex,
        standardisation,
        provenance,
    )
    print(f"created {summary_path.relative_to(ROOT)}")
    print("Overseas regional grouping audit completed.")


if __name__ == "__main__":
    main()
